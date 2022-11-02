from openpyxl import load_workbook

# Load the file into an object
wb_in = load_workbook('input.xlsx')

# Variables and Constants 
ROW_START = 13
ROW_END = 40
teachers = []

# Read the file into a list of dictionary
# for sheet in wb_in:
for sheet in wb_in:
    ws_in = sheet

    name_cell = [3, 10]
    current_column = 2

    for i in range(3):
        new_teacher = {} # create a new teacher dictionary
        new_teacher['name'] = ws_in.cell(row=name_cell[0], column=name_cell[1]).value # enter the name
        new_teacher['am'] = [] # create the in and out in AM list
        new_teacher['pm'] = [] # create the in and out in PM list
        for column in range(0, 4, 2): # no constant necessary here, just the interval
            for row in range(ROW_START, ROW_END):
                new_teacher['am'].append(ws_in.cell(row=row, column = current_column + column).value) # append the am
        current_column += 5
        for column in range(0, 4, 2):
            for row in range(ROW_START, ROW_END):
                new_teacher['pm'].append(ws_in.cell(row=row, column = current_column + column).value) # append the pm
        teachers.append(new_teacher)
        current_column += 9
        name_cell[1] += 15  

for teacher in teachers:
    print(teacher['name'])

