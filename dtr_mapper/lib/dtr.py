from openpyxl import load_workbook

wb_in = load_workbook('src\\input.xlsx')
wb_out = load_workbook('src\\dtr.xlsx')

wb_out_len = len(wb_out.sheetnames)
wb_in_len = len(wb_in.sheetnames)

ws_in = wb_in.active
ws_out = wb_out["Table 1"]
# source = wb_out.active

# # Create the other Worksheets
# if wb_out_len != wb_in_len:
#     for i in range(wb_in_len - 1):
#         wb_out.copy_worksheet(source).title = f'Table {i + 2}'

# In and Out Columns
current = 2 # current column

am = []
pm = []

# IN_AM = B; OUT_AM = C; IN_PM = G; OUT_PM = I
for column in range(0, 4, 2):
    for row in range(13, 40):
        am.append(ws_in.cell(row=row, column=current + column))

current += 5

for column in range(0, 4, 2):
    for row in range(13, 40):
        pm.append(ws_in.cell(row=row, column=current + column))

# for i in range(len(am)):
#     print(am[i].value)

# print()

# for i in pm:
#     print(i.value)

c = 0
for column in range(2, 4):
    for row in range(10, 37):
        ws_out.cell(row=row, column=column, value=am[c].value)
        c += 1

wb_out.save('output.xlsx')