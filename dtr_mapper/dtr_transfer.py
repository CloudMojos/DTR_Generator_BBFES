from openpyxl import load_workbook
import math

# Load the file into an object
wb_in = load_workbook('src\\input.xlsx')
wb_out = load_workbook('src\\dtr.xlsx')

# Variables and Constants 
ROW_START = 13
ROW_END = 40
PRINCIPAL = 'Jocelyn A. San Diego'
COLUMN_PRINCIPAL = [1, 10]
teachers = []

# Read the file into a list of dictionary
# for sheet in wb_in:
for sheet in wb_in:
    ws_in = sheet

    name_cell = [3, 10]
    current_column = 2

    for i in range(3): # 3 because there are three slots in each worksheet
        new_teacher = {} # create a new teacher dictionary
        new_teacher['name'] = ws_in.cell(row=name_cell[0], column=name_cell[1]).value # enter the name
        new_teacher['am'] = [] # create the in and out in AM list
        new_teacher['pm'] = [] # create the in and out in PM list
        for column in range(0, 4, 2): # no constant necessary here, just the interval
            for row in range(ROW_START, ROW_END):
                new_teacher['am'].append(ws_in.cell(row=row, column = current_column + column).value) # append the am
        current_column += 5
        for column in range(0, 4, 2):
            for row in range(ROW_START, ROW_END):
                new_teacher['pm'].append(ws_in.cell(row=row, column = current_column + column).value) # append the pm
        teachers.append(new_teacher)
        current_column += 10
        name_cell[1] += 15  

ws_out = wb_out.active
number_of_teachers = len(teachers)
wb_out_len = len(wb_out.sheetnames)
supposed_number_of_worksheets = math.ceil(number_of_teachers)

print(teachers[0]['name'])
print(teachers[0]['am'])
print(teachers[0]['am'][2])

# Create the other Worksheets
if wb_out_len != supposed_number_of_worksheets:
    for i in range(supposed_number_of_worksheets - 1):
        wb_out.copy_worksheet(ws_out).title = f'Table {i + 2}'

# Add the info to the DTR, start in non-template worksheet
sheets = wb_out.sheetnames
sheets.pop(0)
current_teacher = 0
for sheet in sheets:
    if current_teacher >= len(teachers):
        break
    ws_out = wb_out[sheet]
    name_cell = [3, 2]
    current_column = 2
    for dtr in range(2): # 2 because there are two slots in each worksheet
        ws_out.cell(row=name_cell[0], column=name_cell[1], value=teachers[current_teacher]['name'])
        # AM
        i = 0
        for row in range(10, 38):
            if i < len(teachers[0]['am']):
                ws_out.cell(row=row, column=current_column, value=teachers[current_teacher]['am'][i])
                i += 1
        current_column += 1
        for row in range(10, 38):
            if i < len(teachers[0]['pm']):
                ws_out.cell(row = row, column=current_column, value=teachers[current_teacher]['am'][i])
                i += 1
        # PM
        current_column += 1
        i = 0
        for row in range(10, 38):
            if i < len(teachers[0]['am']):
                ws_out.cell(row=row, column=current_column, value=teachers[current_teacher]['pm'][i])
                i += 1
        current_column += 1
        for row in range(10, 38):
            if i < len(teachers[0]['pm']):
                ws_out.cell(row = row, column=current_column, value=teachers[current_teacher]['pm'][i])
                i += 1
        ws_out.cell(row=44, column=COLUMN_PRINCIPAL[dtr], value=PRINCIPAL)
        current_column = 11
        name_cell[1] = 11
    current_teacher += 1
    

wb_out.save('output.xlsx')